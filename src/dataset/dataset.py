from torch.utils.data import Dataset
from PIL import Image
import openpyxl
import os

class VesselDataset(Dataset):
    def __init__(self, image_paths, label_paths, transform=None, image_transform=None, label_transform=None, quality_file=None):
        self.image_paths = image_paths
        self.label_paths = label_paths
        self.transform = transform
        self.image_transform = image_transform if image_transform is not None else transform
        self.label_transform = label_transform if label_transform is not None else transform

        # Charger le fichier Quality_Assessment.xlsx
        # Clé : (Disease, Number) -> {"IC": int, "Blur": int, "LC": int}
        self.quality_map = {}
        if quality_file is not None and os.path.exists(quality_file):
            wb = openpyxl.load_workbook(quality_file, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    disease, number, ic, blur, lc, *_ = row
                    self.quality_map[(str(disease), int(number))] = {
                        "IC": int(ic),
                        "Blur": int(blur),
                        "LC": int(lc),
                    }
            wb.close()

    def __len__(self):
        return len(self.image_paths)

    def __getitem__(self, idx):
        image_path = self.image_paths[idx]
        image = self.load_img(image_path)

        label_path = self.label_paths[idx]
        label = self.load_label(label_path)

        if self.image_transform:
            image = self.image_transform(image)

        if self.label_transform:
            label = self.label_transform(label)
        
        desease = self.find_desease_from_name(image_path)
        
        quality = self.find_quality_from_quality_assessment(image_path)
        
        return image, label, desease, quality

    def load_img(self, file_path):
        return Image.open(file_path)

    def load_label(self, file_path):
        # In grayscale (1 canal)
        return Image.open(file_path).convert('L')

    def find_desease_from_name(self, file_name):
        # On enlève l'extension en .png et on regarde le dernier caractère (N : normal, A : age-related macular degeneration, G : glaucoma, D : diabetic retinopathy)
        letter = file_name.replace(".png", "")[-1]
        if letter not in ["N", "A", "G", "D"]:
            raise ValueError(f"La lettre {letter} n'est pas reconue")
        else:
            return letter
        
    def find_quality_from_quality_assessment(self, file_name):
        # IC : illumination and color
        # Blur : blur
        # LC : low contrast
        # Extrait le numéro et la maladie depuis le nom de fichier (ex: "29_A.png" -> number=29, disease="A")
        basename = os.path.basename(file_name).replace(".png", "")
        parts = basename.split("_")
        number = int(parts[0])
        disease = parts[1]
        
        key = (disease, number)
        if key in self.quality_map:
            return self.quality_map[key]  # {"IC": int, "Blur": int, "LC": int}
        else:
            # Valeur par défaut pour éviter les erreurs de collation du DataLoader
            return {"IC": -1, "Blur": -1, "LC": -1}